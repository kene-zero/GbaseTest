#!/bin/python3
# -*- coding:utf-8 -*-

import string
import os
from xmindparser import xmind_to_dict
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, PatternFill, Side, Alignment


class XmindToCase(object):

    def __init__(self, XMindPath=None, CaseFilePath=None):

        self.root_path = os.path.dirname(os.path.dirname(__file__))
        self.files_path = os.path.join(self.root_path, "files")
        self.xmind_path = XMindPath if XMindPath else "D:\\File\\脚本文件\\测试设计"
        self.testCase_path = CaseFilePath if CaseFilePath else "D:\\File\\脚本文件\\测试用例"
        self.xmind_dict_list = []
        self.func_name_list = []

    def xmind_dict(self):
        xmind_file_list = os.listdir(self.xmind_path)
        print(xmind_file_list)
        for xmind_file in xmind_file_list:
            xmind_dict = xmind_to_dict(os.path.join(self.xmind_path, xmind_file))
            self.xmind_dict_list.append(xmind_dict)
            self.func_name_list.append(xmind_file[0:-6])

    def dict_to_case(self, xmind_dict):
        case_list = []
        history_info = xmind_dict[0]['topic']["topics"]
        history_name = xmind_dict[0]['topic']["title"]
        print(history_name)
        _info = {}
        for info in history_info:  # 第一层
            if info['title'] == "需求说明":
                for tips in info['topics']:
                    key = tips["title"]
                    _info[key] = tips["topics"][0]['title']
            else:
                case_dict = info['topics']
                for func in case_dict:  # 第二层
                    for test_type in func["topics"]:
                        for test_title in test_type["topics"]:
                            steps = []
                            expect = []
                            for step in test_title["topics"]:
                                steps.append(step["title"])
                                expect.append(step["topics"][0]['title']) if step.get('topics') else expect.append("NA")
                            case = {"test_func": func['title'],
                                    "test_type": test_type['title'],
                                    "test_title": test_title["title"],
                                    "test_step": steps,
                                    'test_expect': expect,
                                    'database': _info["集群类型"],
                                    'ID': _info["用例编号"]}
                            case_list.append(case)
        return case_list

    def generate_excel(self):
        wb = Workbook()
        self.xmind_dict()
        index = 0
        for xmind_dict, func_name in zip(self.xmind_dict_list, self.func_name_list):
            case_list = self.dict_to_case(xmind_dict)
            wb.create_sheet(index=index, title=func_name)
            index += 1
            sheet = wb[func_name]
            num = 1
            sheet[f'A{str(num)}'] = "Test linkID"
            sheet[f'B{str(num)}'] = "用例编号"
            sheet[f'C{str(num)}'] = "测试功能"
            sheet[f'D{str(num)}'] = "测试类别"
            sheet[f'E{str(num)}'] = "测试标题"
            sheet[f'F{str(num)}'] = "测试环境"
            sheet[f'G{str(num)}'] = "前置条件"
            sheet[f'H{str(num)}'] = "测试步骤"
            sheet[f'I{str(num)}'] = "检查项"
            sheet[f'J{str(num)}'] = "优先级"
            sheet[f'K{str(num)}'] = "是否自动化"
            sheet[f'L{str(num)}'] = "用例作者"
            for case in case_list:
                num += 1
                sheet[f'A{str(num)}'] = " "
                sheet[f'B{str(num)}'] = case["ID"] + str(num + 1)
                sheet[f'C{str(num)}'] = case["test_func"]
                sheet[f'D{str(num)}'] = case["test_type"]
                sheet[f'E{str(num)}'] = case["test_title"]
                sheet[f'F{str(num)}'] = case["database"]
                sheet[f'G{str(num)}'] = """1.集群已部署
2.集群状态正常"""
                step = ""
                expect = ''
                n = 1
                for s, e in zip(case["test_step"], case["test_expect"]):
                    s_f = "步骤" + str(n) + ': ' + s + '\r\n'
                    e_f = "步骤" + str(n) + ': ' + e + '\r\n'
                    step += s_f
                    expect += e_f
                    n += 1
                sheet[f'H{str(num)}'] = step.strip()
                sheet[f'I{str(num)}'] = expect.strip()
                sheet[f'J{str(num)}'] = "★★★"
                sheet[f'K{str(num)}'] = "否"
                sheet[f'L{str(num)}'] = "ShiShaoHua"
            excel_file_path = os.path.join("D:\\File\\测试文件\\测试用例\\测试用例.xlsx")
            wb.save(excel_file_path)

    def update_style(self):
        wb = load_workbook("D:\\File\\测试文件\\测试用例\\测试用例.xlsx")
        print(wb.sheetnames)
        ws = wb[wb.sheetnames[0]]
        print("row:", ws.max_row, "column:", ws.max_column)
        height_list = ['10', '15', '15', '15', '30', '10', '20', '40', '35', '10', '10', '15']
        for i in range(1, ws.max_row + 1):
            ws.row_dimensions[i].height = 40 if i != 1 else 20
        for i, w in zip(range(1, ws.max_column + 1), height_list):
            ws.column_dimensions[get_column_letter(i)].width = w

        for x in list(string.ascii_uppercase)[0:12]:
            ws[f'{x}1'].fill = PatternFill("solid", fgColor="C0C0C0")
            ws[f'{x}1'].font = Font(color='000000', bold=True)
            ws[f'{x}1'].border = Border(left=Side(style='thin'), bottom=Side(style='thin'), right=Side(style='thin'),
                                        top=Side(style='thin'))
            ws[f'{x}1'].alignment = Alignment(horizontal='left')
            for y in range(2, ws.max_row + 1):
                ws[f'{x}{y}'].border = Border(left=Side(style='thin'), bottom=Side(style='thin'),
                                              right=Side(style='thin'),
                                              top=Side(style='thin'))
                ws[f'{x}{y}'].alignment = Alignment(wrap_text=True, horizontal='left')

        wb.save("D:\\File\\脚本文件\\测试用例\\测试用例.xlsx")


if __name__ == '__main__':
    tool = XmindToCase()
    tool.generate_excel()
    tool.update_style()
